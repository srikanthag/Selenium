import openpyxl

Path = r"C:\Users\hp\Desktop\IT\Testing\Selenium_Class\Data_driven_testing\Data.xlsx"

workbook = openpyxl.load_workbook(Path)

sheet = workbook.get_sheet_by_name("Sheet1")    ## if u have multiple sheets
sheet = workbook.active


Rows = sheet.max_row
print(Rows)

Columns = sheet.max_column
print(Columns)

for r in range(1,Rows+1):
    for c in range(1, Columns+1):
        print(sheet.cell(row=r,column=c).value, end="    " )
    print()

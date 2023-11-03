import openpyxl

def getRowCount(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)
    return (sheet.max_row)


def getColumnCount(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)
    return (sheet.max_column)


def ReadData(file, sheetname, rowno, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)
    return sheet.cell(row=rowno, column=columnno).value



# def WriteData(file, sheetname, rowno, columnno,data):
#     workbook = openpyxl.load_workbook(file)
#     sheet = openpyxl.get_sheet_by_name(sheetname)
#     return sheet.cell(row=rowno, column=columnno).value = data
#     workbook.save(file)


from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
import time

# Load Excel file
wb = load_workbook('test_data.xlsx')
sheet = wb.active

# Initialize WebDriver
driver = webdriver.Chrome(executable_path='/path/to/chromedriver')


# Function to read data from Excel
def read_excel_data():
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(row)
    return data


# Test case function
def test_login(username, password):
    # Navigate to login page (example URL)
    driver.get('https://example.com/login')

    # Find username and password fields
    username_field = driver.find_element(By.NAME, 'username')
    password_field = driver.find_element(By.NAME, 'password')

    # Clear any existing values
    username_field.clear()
    password_field.clear()

    # Enter the test data
    username_field.send_keys(username)
    password_field.send_keys(password)

    # Submit the form
    password_field.send_keys(Keys.RETURN)

    # Wait for the page to load (adjust timing as necessary)
    time.sleep(3)

    # Check for a successful login (this could vary depending on your application)
    try:
        # For example, checking if the logout button is present
        driver.find_element(By.ID, 'logout')
        print(f"Login successful for {username}")
    except:
        print(f"Login failed for {username}")


# Main function
def run_tests():
    test_data = read_excel_data()
    for data in test_data:
        username, password = data
        test_login(username, password)


# Run the tests
run_tests()

# Close the browser after testing
driver.quit()
